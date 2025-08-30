import os
import json
from datetime import datetime

import requests
from settings import settings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# Global variable to store the backup filename for the session
_session_backup_file = None
# Global array to store all rows
_session_rows = []

def write_row_to_backup(row):
    global _session_backup_file, _session_rows
    
    # Create backup directory if it doesn't exist
    backup_dir = "./data/backup"
    os.makedirs(backup_dir, exist_ok=True)
    
    # Create backup file only once per session
    if _session_backup_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        _session_backup_file = f"{backup_dir}/backup_{timestamp}.json"
        # Initialize the JSON file with an empty array
        with open(_session_backup_file, "w") as f:
            json.dump([], f)
    
    try:
        # Try to append as JSON to the array
        _session_rows.append(row)
        # Write the entire updated array as JSON
        with open(_session_backup_file, "w") as f:
            json.dump(_session_rows, f, indent=2)
    except (TypeError, ValueError) as e:
        # If JSON fails, create a separate text file for this error
        error_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        error_file = f"{backup_dir}/error_{error_timestamp}.txt"
        with open(error_file, "w") as f:
            f.write(f"Error dumping row: {str(e)}\n")
            f.write(f"Row content: {str(row)}\n")

def _get_gender(user_data):
    if not "name" in user_data:
        return "Desconocido"
    name = user_data["name"]
    params = {
        "name": name,
        "key": settings.GENDER_API_KEY
    }
    try:
        response = requests.get(settings.GENDER_URL, params=params, timeout=20)
    except Exception as e:
        print(f"Error getting gender for {name}: {e}")
        return "Desconocido"

    if response.status_code == 200:
        print(f"Gender for {name}: {response.json()}")
        return response.json()["gender"]

def _extract_tweet_data(row):
    """Extract and format tweet data from a row"""
    user_data = row.get("includes", {}).get("users", {})
    user_data = user_data[0] if isinstance(user_data, list) else user_data
    
    pm = row.get("public_metrics", {})
    if not isinstance(pm, dict):
        pm = {}
    
    return {
        "tweet_id": row.get("id", ""),
        "fecha": row.get("created_at", ""),
        "lenguaje": row.get("lang", ""),
        "texto": row.get("text", ""),
        "usuario": user_data.get("username", "") if isinstance(user_data, dict) else "",
        "usuario_nombre": user_data.get("name", "") if isinstance(user_data, dict) else "",
        "usuario_genero": _get_gender(user_data),
        "usuario_verified": user_data.get("verified", "") if isinstance(user_data, dict) else "",
        "usuario_verified_type": user_data.get("verified_type", "") if isinstance(user_data, dict) else "",
        "usuario_ubicacion": user_data.get("location", "") if isinstance(user_data, dict) else "",
        "seguidores": user_data.get("public_metrics", {}).get("followers_count", "") if isinstance(user_data, dict) and isinstance(user_data.get("public_metrics"), dict) else "",
        "siguiendo": user_data.get("public_metrics", {}).get("following_count", "") if isinstance(user_data, dict) and isinstance(user_data.get("public_metrics"), dict) else "",
        "tweets": user_data.get("public_metrics", {}).get("tweet_count", "") if isinstance(user_data, dict) and isinstance(user_data.get("public_metrics"), dict) else "",
        "retweets": pm.get("retweet_count", "") if isinstance(pm, dict) else "",
        "replies": pm.get("reply_count", "") if isinstance(pm, dict) else "",
        "likes": pm.get("like_count", "") if isinstance(pm, dict) else "",
        "quotes": pm.get("quote_count", "") if isinstance(pm, dict) else "",
        "user_dump": json.dumps(user_data, ensure_ascii=False, indent=2) if isinstance(user_data, dict) else "{}",
        "public_metrics_dump": json.dumps(pm, ensure_ascii=False, indent=2) if isinstance(pm, dict) else "{}",
        "tweet_dump": json.dumps(row, ensure_ascii=False, indent=2)
    }

def _group_data_by_language(rows):
    """Group tweet data by language"""
    language_groups = {}
    
    for row in rows:
        tweet_data = _extract_tweet_data(row)
        language = tweet_data["lenguaje"] or "unknown"
        
        if language not in language_groups:
            language_groups[language] = []
        
        language_groups[language].append(tweet_data)
    
    return language_groups

def _create_excel_workbook():
    """Create a new Excel workbook with basic styling"""
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    return wb

def _add_worksheet_with_data(wb, language, data_rows):
    """Add a worksheet for a specific language with formatted data"""
    # Create worksheet name (Excel has restrictions on sheet names)
    sheet_name = _sanitize_sheet_name(language)
    
    # Create worksheet
    ws = wb.create_sheet(title=sheet_name)
    
    # Define headers
    headers = [
        "Tweet ID", "Fecha", "Lenguaje", "Texto", "Usuario", "Usuario Nombre", "Usuario Genero",
        "Verificado", "Tipo de Verificación", "Ubicación", "Seguidores", "Siguiendo", 
        "Tweets", "Retweets", "Replies", "Likes", "Quotes",
        "User Dump", "Public Metrics Dump", "Tweet Dump"
    ]
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data.values(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    _auto_adjust_columns(ws, data_rows)
    
    return ws

def _sanitize_sheet_name(language):
    """Sanitize language name for Excel sheet naming"""
    # Excel sheet names cannot exceed 31 characters and cannot contain certain characters
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    sheet_name = language
    
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '_')
    
    # Truncate if too long
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    
    # Ensure it's not empty
    if not sheet_name:
        sheet_name = "unknown"
    
    return sheet_name

def _auto_adjust_columns(ws, data_rows):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set a reasonable width (min 10, max 50)
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def _save_excel_file(wb, filename=None):
    """Save the Excel workbook to file"""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{settings.OUTPUT_DIR}/tweets_{timestamp}.xlsx"
    
    # Create data directory if it doesn't exist
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    try:
        wb.save(filename)
        logging.info(f"Excel file saved successfully: {filename}")
        return filename
    except Exception as e:
        logging.error(f"Error saving Excel file: {e}")
        raise

def write_rows_to_xlmns(rows, output_filename=None):
    """
    Write tweet rows to Excel file with separate tabs per language
    
    Args:
        rows: List of tweet data dictionaries
        output_filename: Optional custom filename for the Excel file
    
    Returns:
        str: Path to the created Excel file
    """
    try:
        # Group data by language
        language_groups = _group_data_by_language(rows)
        
        if not language_groups:
            logging.warning("No data to write to Excel")
            return None
        
        # Create Excel workbook
        wb = _create_excel_workbook()
        
        # Add worksheet for each language
        for language, data_rows in language_groups.items():
            try:
                _add_worksheet_with_data(wb, language, data_rows)
                logging.info(f"Added worksheet for language: {language} with {len(data_rows)} rows")
            except Exception as e:
                logging.error(f"Error creating worksheet for language {language}: {e}")
                continue
        
        # Save the workbook
        filename = _save_excel_file(wb, output_filename)
        
        return filename
        
    except Exception as e:
        logging.error(f"Error in write_rows_to_xlmns: {e}")
        raise

def write_single_row_to_excel(row, existing_filename=None):
    """
    Write a single tweet row to an existing Excel file or create a new one
    
    Args:
        row: Single tweet data dictionary
        existing_filename: Optional path to existing Excel file to append to
    
    Returns:
        str: Path to the Excel file
    """
    try:
        if existing_filename and os.path.exists(existing_filename):
            # Load existing workbook
            from openpyxl import load_workbook
            wb = load_workbook(existing_filename)
        else:
            # Create new workbook
            wb = _create_excel_workbook()
        
        # Extract data
        tweet_data = _extract_tweet_data(row)
        language = tweet_data["lenguaje"] or "unknown"
        
        # Check if language sheet exists
        sheet_name = _sanitize_sheet_name(language)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Create new sheet for this language
            ws = _add_worksheet_with_data(wb, language, [tweet_data])
            return _save_excel_file(wb, existing_filename)
        
        # Add new row
        next_row = ws.max_row + 1
        for col_idx, value in enumerate(tweet_data.values(), 1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Auto-adjust columns
        _auto_adjust_columns(ws, [tweet_data])
        
        # Save
        return _save_excel_file(wb, existing_filename)
        
    except Exception as e:
        logging.error(f"Error in write_single_row_to_excel: {e}")
        raise


